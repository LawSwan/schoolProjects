import openpyxl

# Load the Excel2 workbook
book = openpyxl.load_workbook("SecondBook.xlsx")
sheet = book.active

# Read the data and modify it
values = []
for row in sheet.iter_rows():
    for cell in row:
        value = cell.value
        if cell.column == 2:
            value = value * 10
        values.append(value)

# Close the original workbook
book.close()

# Create a new workbook and write the modified data
new_book = openpyxl.Workbook()
new_sheet = new_book.active

index = 0
rw = 1
col = 1
for val in values:
    new_sheet.cell(row=rw, column=col, value=val)
    index += 1
    col += 1
    if index == 3:
        rw += 1
        col = 1
        index = 0

# Save the new workbook
new_book.save("ThirdBook.xlsx")